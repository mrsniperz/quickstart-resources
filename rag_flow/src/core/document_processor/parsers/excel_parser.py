"""
模块名称: excel_parser
功能描述: Excel文档解析器，基于openpyxl实现Excel文档的数据提取、工作表处理和批量处理优化
创建日期: 2024-01-15
作者: Sniperz
版本: v1.0.0
"""

import logging
from typing import Dict, List, Optional, Any, Union
from pathlib import Path
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..extractors.metadata_extractor import MetadataExtractor


@dataclass
class ExcelParseResult:
    """Excel解析结果数据类"""
    text_content: str
    metadata: Dict[str, Any]
    worksheets: List[Dict[str, Any]]
    tables: List[Dict[str, Any]]
    structure_info: Dict[str, Any]


class ExcelParser:
    """
    Excel文档处理器
    
    基于openpyxl实现Excel文档的全面解析，包括：
    - 多工作表数据提取
    - 表格数据处理
    - 公式和格式保持
    - 批量处理优化
    - 元数据提取
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化Excel解析器
        
        Args:
            config (dict, optional): 配置参数
                - read_only (bool): 只读模式，默认True
                - data_only (bool): 只读取数据值，默认True
                - extract_formulas (bool): 是否提取公式，默认False
                - max_rows (int): 最大读取行数，默认None
                - max_cols (int): 最大读取列数，默认None
        """
        self.config = config or {}
        self.logger = logging.getLogger(__name__)
        
        # 初始化提取器
        self.metadata_extractor = MetadataExtractor()
        
        # 配置参数
        self.read_only = self.config.get('read_only', True)
        self.data_only = self.config.get('data_only', True)
        self.extract_formulas = self.config.get('extract_formulas', False)
        self.max_rows = self.config.get('max_rows', None)
        self.max_cols = self.config.get('max_cols', None)
        
    def parse(self, file_path: str) -> ExcelParseResult:
        """
        解析Excel文档
        
        Args:
            file_path (str): Excel文件路径
            
        Returns:
            ExcelParseResult: 解析结果
            
        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式不支持
            Exception: 解析过程中的其他错误
        """
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"Excel文件不存在: {file_path}")
                
            if file_path.suffix.lower() not in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
                raise ValueError(f"不支持的文件格式: {file_path.suffix}")
                
            self.logger.info(f"开始解析Excel文档: {file_path}")
            
            # 打开工作簿
            wb = load_workbook(
                filename=str(file_path),
                read_only=self.read_only,
                data_only=self.data_only
            )
            
            try:
                # 提取元数据
                metadata = self._extract_metadata(wb, file_path)
                
                # 提取工作表数据
                worksheets = self._extract_worksheets(wb)
                
                # 提取表格数据
                tables = self._extract_tables(wb)
                
                # 生成文本内容
                text_content = self._generate_text_content(worksheets)
                
                # 分析文档结构
                structure_info = self._analyze_structure(wb, worksheets)
                
                result = ExcelParseResult(
                    text_content=text_content,
                    metadata=metadata,
                    worksheets=worksheets,
                    tables=tables,
                    structure_info=structure_info
                )
                
                self.logger.info(f"Excel解析完成: {len(worksheets)}个工作表, {len(tables)}个表格")
                return result
                
            finally:
                if self.read_only:
                    wb.close()
                    
        except Exception as e:
            self.logger.error(f"Excel解析失败: {e}")
            raise
    
    def _extract_metadata(self, wb: Workbook, file_path: Path) -> Dict[str, Any]:
        """
        提取工作簿元数据
        
        Args:
            wb: openpyxl工作簿对象
            file_path: 文件路径
            
        Returns:
            dict: 元数据信息
        """
        try:
            props = wb.properties
            
            metadata = {
                'title': props.title or '',
                'creator': props.creator or '',
                'subject': props.subject or '',
                'description': props.description or '',
                'keywords': props.keywords or '',
                'category': props.category or '',
                'comments': props.comments or '',
                'created': props.created,
                'modified': props.modified,
                'last_modified_by': props.lastModifiedBy or '',
                'version': props.version or '',
                'file_path': str(file_path),
                'file_size': file_path.stat().st_size if file_path.exists() else 0,
                'file_extension': file_path.suffix.lower(),
                'worksheet_count': len(wb.worksheets),
                'worksheet_names': wb.sheetnames,
                'is_excel_document': True
            }
            
            return metadata
            
        except Exception as e:
            self.logger.error(f"元数据提取失败: {e}")
            return {'file_path': str(file_path), 'is_excel_document': True}
    
    def _extract_worksheets(self, wb: Workbook) -> List[Dict[str, Any]]:
        """
        提取所有工作表数据
        
        Args:
            wb: openpyxl工作簿对象
            
        Returns:
            list: 工作表信息列表
        """
        try:
            worksheets = []
            
            for ws_idx, ws in enumerate(wb.worksheets):
                try:
                    ws_info = self._extract_worksheet_data(ws, ws_idx)
                    if ws_info:
                        worksheets.append(ws_info)
                        
                except Exception as e:
                    self.logger.warning(f"工作表'{ws.title}'提取失败: {e}")
                    continue
            
            return worksheets
            
        except Exception as e:
            self.logger.error(f"工作表提取失败: {e}")
            return []
    
    def _extract_worksheet_data(self, ws: Worksheet, ws_idx: int) -> Optional[Dict[str, Any]]:
        """
        提取单个工作表数据
        
        Args:
            ws: openpyxl工作表对象
            ws_idx: 工作表索引
            
        Returns:
            dict: 工作表信息，如果工作表为空则返回None
        """
        try:
            # 确定数据范围
            if ws.max_row == 1 and ws.max_column == 1:
                # 检查是否真的为空
                if not ws.cell(1, 1).value:
                    return None
            
            max_row = min(ws.max_row, self.max_rows) if self.max_rows else ws.max_row
            max_col = min(ws.max_column, self.max_cols) if self.max_cols else ws.max_column
            
            # 提取数据
            data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, 
                                  min_col=1, max_col=max_col, 
                                  values_only=True):
                # 转换None为空字符串，其他值转为字符串
                row_data = [str(cell) if cell is not None else '' for cell in row]
                data.append(row_data)
            
            # 移除完全空白的行
            data = [row for row in data if any(cell.strip() for cell in row)]
            
            if not data:
                return None
            
            ws_info = {
                'index': ws_idx,
                'name': ws.title,
                'data': data,
                'rows': len(data),
                'columns': len(data[0]) if data else 0,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'has_tables': len(ws.tables) > 0,
                'table_names': [table.name for table in ws.tables]
            }
            
            return ws_info
            
        except Exception as e:
            self.logger.error(f"工作表数据提取失败: {e}")
            return None
    
    def _extract_tables(self, wb: Workbook) -> List[Dict[str, Any]]:
        """
        提取工作簿中的所有表格
        
        Args:
            wb: openpyxl工作簿对象
            
        Returns:
            list: 表格信息列表
        """
        try:
            all_tables = []
            
            for ws in wb.worksheets:
                for table in ws.tables.values():
                    try:
                        # 获取表格范围
                        table_range = table.ref
                        
                        # 提取表格数据
                        table_data = []
                        for row in ws[table_range]:
                            row_data = [str(cell.value) if cell.value is not None else '' 
                                      for cell in row]
                            table_data.append(row_data)
                        
                        if table_data:
                            table_info = {
                                'name': table.name,
                                'worksheet': ws.title,
                                'range': table_range,
                                'data': table_data,
                                'rows': len(table_data),
                                'columns': len(table_data[0]) if table_data else 0,
                                'has_headers': table.tableStyleInfo.showFirstColumn if table.tableStyleInfo else False
                            }
                            all_tables.append(table_info)
                            
                    except Exception as e:
                        self.logger.warning(f"表格'{table.name}'提取失败: {e}")
                        continue
            
            return all_tables
            
        except Exception as e:
            self.logger.error(f"表格提取失败: {e}")
            return []
    
    def _generate_text_content(self, worksheets: List[Dict[str, Any]]) -> str:
        """
        从工作表数据生成文本内容
        
        Args:
            worksheets: 工作表信息列表
            
        Returns:
            str: 生成的文本内容
        """
        try:
            text_parts = []
            
            for ws in worksheets:
                ws_text = [f"=== 工作表: {ws['name']} ==="]
                
                for row_idx, row in enumerate(ws['data']):
                    # 跳过完全空白的行
                    if any(cell.strip() for cell in row):
                        row_text = '\t'.join(row)
                        ws_text.append(f"第{row_idx + 1}行: {row_text}")
                
                text_parts.append('\n'.join(ws_text))
            
            return '\n\n'.join(text_parts)
            
        except Exception as e:
            self.logger.error(f"文本内容生成失败: {e}")
            return ""
    
    def _analyze_structure(self, wb: Workbook, worksheets: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        分析工作簿结构
        
        Args:
            wb: openpyxl工作簿对象
            worksheets: 工作表信息列表
            
        Returns:
            dict: 结构分析信息
        """
        try:
            structure = {
                'total_worksheets': len(worksheets),
                'total_tables': sum(len(ws.get('table_names', [])) for ws in worksheets),
                'total_rows': sum(ws.get('rows', 0) for ws in worksheets),
                'total_columns': max(ws.get('columns', 0) for ws in worksheets) if worksheets else 0,
                'worksheet_summary': [],
                'largest_worksheet': None,
                'has_named_ranges': len(wb.defined_names) > 0,
                'named_ranges': list(wb.defined_names.definedName) if hasattr(wb.defined_names, 'definedName') else []
            }
            
            # 工作表摘要
            max_rows = 0
            largest_ws = None
            
            for ws in worksheets:
                ws_summary = {
                    'name': ws['name'],
                    'rows': ws['rows'],
                    'columns': ws['columns'],
                    'has_tables': ws.get('has_tables', False),
                    'table_count': len(ws.get('table_names', []))
                }
                structure['worksheet_summary'].append(ws_summary)
                
                if ws['rows'] > max_rows:
                    max_rows = ws['rows']
                    largest_ws = ws['name']
            
            structure['largest_worksheet'] = largest_ws
            
            return structure
            
        except Exception as e:
            self.logger.error(f"结构分析失败: {e}")
            return {}
    
    def extract_worksheet_data(self, file_path: str, worksheet_name: str) -> List[List[str]]:
        """
        提取指定工作表的数据
        
        Args:
            file_path (str): Excel文件路径
            worksheet_name (str): 工作表名称
            
        Returns:
            list: 工作表数据
            
        Raises:
            ValueError: 工作表不存在
        """
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            
            try:
                if worksheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表不存在: {worksheet_name}")
                
                ws = wb[worksheet_name]
                ws_info = self._extract_worksheet_data(ws, 0)
                
                return ws_info['data'] if ws_info else []
                
            finally:
                wb.close()
                
        except Exception as e:
            self.logger.error(f"工作表数据提取失败: {e}")
            raise
    
    def get_supported_formats(self) -> List[str]:
        """
        获取支持的文件格式
        
        Returns:
            list: 支持的文件扩展名列表
        """
        return ['.xlsx', '.xlsm', '.xltx', '.xltm']
